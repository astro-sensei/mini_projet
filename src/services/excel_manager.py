import openpyxl
from openpyxl import Workbook
from src.services.utils import FormatInvalideError

class ExcelManager:
    """Gère les imports et exports Excel."""

    def exporter_produits(self, filepath, produits):
        """
        Exporte la liste des produits vers un fichier Excel.
        produits: dict ou list d'objets Produit
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Produits"
        
        # En-têtes
        ws.append(["Reference", "Nom", "Categorie", "Quantite", "Prix"])
        
        # Données
        # produits est attendu comme une liste d'objets Produit
        for p in produits:
            ws.append([p.reference, p.nom, p.categorie, p.quantite, p.prix])
            
        wb.save(filepath)

    def exporter_ventes(self, filepath, ventes):
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventes"
        
        ws.append(["Reference Produit", "Quantite", "Date", "Total"])
        
        for v in ventes:
            # On suppose que v.date_vente est un datetime ou str
            date_str = str(v.date_vente)
            # On ne connait pas le prix unitaire ici sans le produit, 
            # mais la méthode appelante devra gérer le total si besoin.
            # Pour l'instant on exporte les données brutes de la vente.
            ws.append([v.reference_produit, v.quantite, date_str])
            
        wb.save(filepath)

    def importer_produits(self, filepath):
        """Retourne une liste de dictionnaires pour création de produits."""
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            
            produits_importes = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                # On attend: Nom, Categorie, Quantite, Prix
                # L'ID ne doit PAS être importé, il est généré par l'app.
                # Si le fichier contient une colonne ID, on l'ignore ou on l'utilise pour verification?
                # Selon la doc : "L'ID est généré automatiquement".
                # Donc on suppose un format d'import : Nom, Categorie, Quantite, Prix
                if not row or row[0] is None:
                    continue
                    
                if len(row) < 4:
                    continue 

                produits_importes.append({
                    "nom": str(row[0]),
                    "categorie": str(row[1]),
                    "quantite": row[2],
                    "prix": row[3]
                })
            return produits_importes
            
        except FileNotFoundError:
            raise FormatInvalideError("Fichier Excel introuvable.")
        except Exception as e:
            raise FormatInvalideError(f"Erreur import Excel: {e}")